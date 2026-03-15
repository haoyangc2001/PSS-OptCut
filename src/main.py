from Pamas_generator import Generator
import numpy as np
import pandas as pd
from pathlib import Path
from Solver_builder import GRBModel
from sample_generator import TrainSetGenerator
from gurobipy import quicksum
import datetime
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
RESULT_CSV_PATH = DATA_DIR / "result.csv"


def sample_instance_size(times):
    """
    按实验阶段生成实例规模。
    前 300 轮保持原始随机区间，后 100 轮进入递增压测段。
    """
    if times <= 300:
        num_customer = np.random.randint(5, 10)
        num_service = np.random.randint(3, 6)
        num_team = np.random.randint(3, 8)
        scale_phase = 'baseline_300'
        scale_stage = 'S1'
    elif times <= 325:
        num_customer = np.random.randint(10, 13)
        num_service = np.random.randint(5, 7)
        num_team = np.random.randint(6, 9)
        scale_phase = 'stress_100'
        scale_stage = 'S2'
    elif times <= 350:
        num_customer = np.random.randint(12, 15)
        num_service = np.random.randint(6, 8)
        num_team = np.random.randint(7, 10)
        scale_phase = 'stress_100'
        scale_stage = 'S3'
    elif times <= 375:
        num_customer = np.random.randint(14, 17)
        num_service = np.random.randint(7, 9)
        num_team = np.random.randint(8, 11)
        scale_phase = 'stress_100'
        scale_stage = 'S4'
    else:
        num_customer = np.random.randint(16, 19)
        num_service = np.random.randint(8, 11)
        num_team = np.random.randint(9, 13)
        scale_phase = 'stress_100'
        scale_stage = 'S5'

    num_product = num_service
    size_signature = f'I{num_customer}_M{num_service}_N{num_product}_K{num_team}'

    return num_customer, num_service, num_product, num_team, scale_phase, scale_stage, size_signature

# 模型目标函数
Obj_Coef = [1, -1, 1, 0]
valid_cut_selection = [1, 1, 1]
# valid cut
valid_cut_pool = {}
valid_cut_pool[
    0] = 'self.model.addConstrs(self.X[i, j, k] + self.X[j, i, k] <= 1 for i in range(self.generator.Num_Customer) for j in range(self.generator.Num_Customer) for k in range(self.generator.Num_Team),name=cut_0)'
valid_cut_pool[
    1] = 'self.model.addConstrs(self.T[i, k] + sum(self.R[i,m]*self.generator.Service_Time[m] for m in range(self.generator.Num_Service)) + self.generator.Transfer_Time[i][j] - self.BigM * (1 - self.X[i, j, k]) <= self.generator.Late_Start_Limit[j] for i in range(self.generator.Num_Customer) for j in range(self.generator.Num_Customer) for k in range(self.generator.Num_Team))'
valid_cut_pool[
    2] = 'self.model.addConstrs(self.generator.Early_Start_Limit[i] + self.generator.Transfer_Time[i][j] + min(self.generator.Service_Time) - self.BigM * (1 - self.X[i, j, k]) <= self.generator.Late_Start_Limit[j] for i in range(self.generator.Num_Customer) for j in range(self.generator.Num_Customer) for k in range(self.generator.Num_Team))'

# ================================================================================
# 启动参数
Num_Customer = 7 # 顾客数
Num_Service = 3 # 服务数
Num_Product = 3 # 产品数
Num_Team = 3 # 服务团队数
random_seed = 6701

"""service"""
service_time_step = 2
service_price_step = 2
service_cost_step = 2
service_time_lb = 1
service_price_lb = 7
service_cost_lb = 1
"""time"""
early_start_lb = 0
early_start_ub = 3
late_start_lb = 4
late_start_ub = 7
transfer_lb = 1
transfer_ub = 2
"""utility"""
utility_lb = 4
utility_step = 1
"""inventory"""
inventory_step = 4
inventory_lb = 4
"""weight"""
weight_tri = 1
weight_gap = 2
"""variable cost"""
variable_cost_lb = 1
variable_cost_ub = 2
"""ser_pro_coe"""
ser_pro_coe = 1  # 服务 ：→ 产品 系数 可以是一个常数  也可以是一个矩阵，即不同的映射关系，要生产不同的关系需要调用gen_ser_pro_coe
ser_pro_coe_lb = 1
ser_pro_coe_ub = 2
"""duration"""
duration = 12

times = 1
TOTAL_TIMES = 400

# np.random.seed(333)
choice = 1

# 参数生成器
gen = Generator()

# 生成样本
while times <= TOTAL_TIMES:
    xl_target_data = []
    if times >= 1:
        Num_Customer, Num_Service, Num_Product, Num_Team, scale_phase, scale_stage, size_signature = sample_instance_size(times)

        random_seed = np.random.randint(low=10, high=10000)

        """service"""
        service_time_step = np.random.randint(2, 6)
        service_price_step = np.random.randint(5, 12)
        service_cost_step = np.random.randint(3, 7)
        service_time_lb = np.random.randint(1, 7)
        service_price_lb = np.random.randint(10, 25)
        service_cost_lb = np.random.randint(1, 7)
        """time"""
        early_start_lb = np.random.randint(0, 5)
        early_start_ub = np.random.randint(5, 10)
        late_start_lb = np.random.randint(10, 15)
        late_start_ub = np.random.randint(15, 35)
        transfer_lb = np.random.randint(1, 3)
        transfer_ub = np.random.randint(3, 6)
        """utility"""
        utility_lb = np.random.randint(4, 12)
        utility_step = np.random.randint(1, 4)
        """inventory"""
        inventory_step = np.random.randint(10, 16)
        inventory_lb = np.random.randint(10, 20)
        """weight"""
        weight_gap = np.random.randint(2, 6)
        """variable cost"""
        variable_cost_lb = np.random.randint(0, 2)
        variable_cost_ub = np.random.randint(2, 3)
        """ser_pro_coe"""
        ser_pro_coe = np.random.randint(1, 20)
        ser_pro_coe_lb = np.random.randint(1, 2)
        ser_pro_coe_ub = np.random.randint(2, 4)
        """duration"""
        duration = np.random.randint(8, 30)

    gen.set_instance(num_customer=Num_Customer, num_service=Num_Service, num_product=Num_Product, num_team=Num_Team,
                     random_seed=random_seed)
    gen.set_params(
        service_time_step=service_time_step, service_price_step=service_price_step, service_cost_step=service_cost_step,
        service_time_lb=service_time_lb, service_price_lb=service_price_lb, service_cost_lb=service_cost_lb,
        early_start_lb=early_start_lb, early_start_ub=early_start_ub,
        late_start_lb=late_start_lb, late_start_ub=late_start_ub,
        transfer_lb=transfer_lb, transfer_ub=transfer_ub,
        utility_lb=utility_lb, utility_step=utility_step,
        inventory_step=inventory_step, inventory_lb=inventory_lb,
        weight_tri=weight_tri, weight_gap=weight_gap,
        variable_cost_lb=variable_cost_lb, variable_cost_ub=variable_cost_ub,
        ser_pro_coe=ser_pro_coe, ser_pro_coe_lb=ser_pro_coe_lb, ser_pro_coe_ub=ser_pro_coe_ub,
        )

    # generate params
    gen.gen_customer_weight(weight_tri=gen.weight_tri, weight_gap=gen.weight_gap)
    gen.gen_service_relations()
    gen.gen_ser_pro_coe()
    gen.gen_customer_relations()
    gen.gen_time_relations()
    gen.gen_dis_matrix()
    gen.gen_cost_relations()

    trigger = 1
    gap = 0

    ml_train_set_gen = TrainSetGenerator()
    # input params
    ml_train_set_gen.update_paras(model_obj_coe=Obj_Coef, model_valid_cut_pool=valid_cut_pool,
                                  model_opt_trigger=trigger, model_gap=gap, generator=gen,
                                  scale_phase=scale_phase, scale_stage=scale_stage,
                                  size_signature=size_signature)
    current_time = datetime.datetime.now().strftime('%Y-%m-%d')
    # filename = str(current_time) + '_' + str(times) +'.xlsx'
    ml_train_set_gen.build_optimize()
    current_df = ml_train_set_gen.pd_to_excel(data=ml_train_set_gen.xl_data)  # 返回排序后的df

    # xl_target_data.append(ml_train_set_gen.xl_data[current_df.iloc[0, 0]])
    # xl_target_data.append(current_df)

    for i in range(len(ml_train_set_gen.valid_cut_selection_permutation)):
        xl_target_data.append(ml_train_set_gen.xl_data[current_df.iloc[i, 0]])

    # 记录结果
    xl_id = []
    xl_valid_selection = []
    xl_opt_trigger = []
    xl_random_seed = []
    xl_gap = []
    xl_scale_phase = []
    xl_scale_stage = []
    xl_size_signature = []

    xl_customer_num = []
    xl_service_num = []
    xl_product_num = []
    xl_team_num = []
    xl_base_num_vars = []
    xl_base_num_constrs = []
    xl_root_num_constrs = []

    xl_service_time = []
    xl_service_price = []
    xl_service_cost = []

    xl_profit = []
    xl_profit_mean = []
    xl_profit_std = []
    xl_profit_high4 = []
    xl_profit_low4 = []

    xl_service_time_lb = []
    xl_service_price_lb = []
    xl_service_cost_lb = []

    xl_service_time_step = []
    xl_service_price_step = []
    xl_service_cost_step = []

    xl_utility = []
    xl_utility_lb = []
    xl_utility_step = []
    xl_inventory = []
    xl_inventory_step = []
    xl_inventory_lb = []

    xl_Ser_Pro_corr = []
    xl_Ser_Pro_coe = []
    xl_Ser_Pro_lb = []
    xl_Ser_Pro_ub = []
    # mean
    xl_Ser_Pro_coe_mean = []
    # std
    xl_Ser_Pro_coe_std = []
    # high 4
    xl_Ser_Pro_coe_high4 = []

    xl_weight = []
    xl_weight_gap = []
    xl_weight_mean = []
    xl_weight_std = []
    xl_weight_high4 = []

    xl_early_start = []
    xl_early_start_lb = []
    xl_early_start_ub = []
    xl_late_start = []
    xl_late_start_lb = []
    xl_late_start_ub = []

    # mean
    xl_early_start_mean = []
    xl_late_start_mean = []
    # std
    xl_early_start_std = []
    xl_late_start_std = []
    # high4
    xl_early_start_high4 = []
    xl_late_start_high4 = []

    xl_transfer_matrix = []
    xl_transfer_lb = []
    xl_transfer_ub = []
    xl_transfer_mean = []
    xl_transfer_std = []
    xl_transfer_high4 = []

    xl_duration = []

    xl_fixed_cost = []
    xl_variable_cost = []
    xl_variable_cost_lb = []
    xl_variable_cost_ub = []
    xl_variable_cost_mean = []
    xl_variable_cost_std = []
    xl_variable_cost_high4 = []

    xl_cut3_percent = []
    # =======  results ======
    xl_obj1 = []  # 总性价比
    xl_obj2 = []  # 总任务开始时间和
    xl_obj3 = []  # 总利润
    xl_obj = []  # 函数目标值

    xl_runtime = []

    for i in range(len(xl_target_data)):
        xl_id.append(xl_target_data[i]["id"])

        xl_valid_selection.append(xl_target_data[i]["valid_selection"])
        xl_opt_trigger.append(xl_target_data[i]["opt_trigger"])
        xl_random_seed.append(xl_target_data[i]["random_seed"])
        xl_gap.append(xl_target_data[i]["gap"])
        xl_scale_phase.append(xl_target_data[i]["scale_phase"])
        xl_scale_stage.append(xl_target_data[i]["scale_stage"])
        xl_size_signature.append(xl_target_data[i]["size_signature"])

        xl_customer_num.append(xl_target_data[i]["customer_num"])
        xl_service_num.append(xl_target_data[i]["service_num"])
        xl_product_num.append(xl_target_data[i]["product_num"])
        xl_team_num.append(xl_target_data[i]["team_num"])
        xl_base_num_vars.append(xl_target_data[i]["base_num_vars"])
        xl_base_num_constrs.append(xl_target_data[i]["base_num_constrs"])
        xl_root_num_constrs.append(xl_target_data[i]["root_num_constrs"])

        xl_service_time.append(xl_target_data[i]["service_time"])
        xl_service_price.append(xl_target_data[i]["service_price"])
        xl_service_cost.append(xl_target_data[i]["service_cost"])

        xl_profit.append(xl_target_data[i]["profit"])
        xl_profit_mean.append(xl_target_data[i]["profit_mean"])
        xl_profit_std.append(xl_target_data[i]["profit_std"])
        xl_profit_high4.append(xl_target_data[i]["profit_high4"])
        xl_profit_low4.append(xl_target_data[i]["profit_low4"])

        xl_service_time_lb.append(xl_target_data[i]["service_time_lb"])
        xl_service_price_lb.append(xl_target_data[i]["service_price_lb"])
        xl_service_cost_lb.append(xl_target_data[i]["service_cost_lb"])

        xl_service_time_step.append(xl_target_data[i]["service_time_step"])
        xl_service_price_step.append(xl_target_data[i]["service_price_step"])
        xl_service_cost_step.append(xl_target_data[i]["service_cost_step"])

        xl_utility.append(xl_target_data[i]["utility"])
        xl_utility_lb.append(xl_target_data[i]["utility_lb"])
        xl_utility_step.append(xl_target_data[i]["utility_step"])
        xl_inventory.append(xl_target_data[i]["inventory"])
        xl_inventory_step.append(xl_target_data[i]["inventory_step"])
        xl_inventory_lb.append(xl_target_data[i]["inventory_lb"])

        xl_Ser_Pro_corr.append(xl_target_data[i]["Ser_Pro_corr"])
        xl_Ser_Pro_coe.append(xl_target_data[i]["Ser_Pro_coe"])
        xl_Ser_Pro_lb.append(xl_target_data[i]["ser_pro_coe_lb"])
        xl_Ser_Pro_ub.append(xl_target_data[i]["ser_pro_coe_ub"])
        # mean
        xl_Ser_Pro_coe_mean.append(xl_target_data[i]["Ser_Pro_coe_mean"])
        # std
        xl_Ser_Pro_coe_std.append(xl_target_data[i]["Ser_Pro_coe_std"])
        # high 4
        xl_Ser_Pro_coe_high4.append(xl_target_data[i]["Ser_Pro_coe_high4"])

        xl_weight.append(xl_target_data[i]["weight"])
        xl_weight_gap.append(xl_target_data[i]["weight_gap"])
        xl_weight_mean.append(xl_target_data[i]["weight_mean"])
        xl_weight_std.append(xl_target_data[i]["weight_std"])
        xl_weight_high4.append(xl_target_data[i]["weight_high4"])

        xl_early_start.append(xl_target_data[i]["early_start"])
        xl_early_start_lb.append(xl_target_data[i]["early_start_lb"])
        xl_early_start_ub.append(xl_target_data[i]["early_start_ub"])
        xl_late_start.append(xl_target_data[i]["late_start"])
        xl_late_start_lb.append(xl_target_data[i]["late_start_lb"])
        xl_late_start_ub.append(xl_target_data[i]["late_start_ub"])

        # mean
        xl_early_start_mean.append(xl_target_data[i]["early_start_mean"])
        xl_late_start_mean.append(xl_target_data[i]["late_start_mean"])
        # std
        xl_early_start_std.append(xl_target_data[i]["early_start_std"])
        xl_late_start_std.append(xl_target_data[i]["late_start_std"])
        # high4
        xl_early_start_high4.append(xl_target_data[i]["early_start_high4"])
        xl_late_start_high4.append(xl_target_data[i]["late_start_high4"])

        xl_transfer_matrix.append(xl_target_data[i]["transfer_matrix"])
        xl_transfer_lb.append(xl_target_data[i]["transfer_lb"])
        xl_transfer_ub.append(xl_target_data[i]["transfer_ub"])
        xl_transfer_mean.append(xl_target_data[i]["transfer_mean"])
        xl_transfer_std.append(xl_target_data[i]["transfer_std"])
        xl_transfer_high4.append(xl_target_data[i]["transfer_high4"])

        xl_duration.append(xl_target_data[i]["duration"])

        xl_fixed_cost.append(xl_target_data[i]["FixedCost"])
        xl_variable_cost.append(xl_target_data[i]["VariableCost"])
        xl_variable_cost_lb.append(xl_target_data[i]["variable_cost_lb"])
        xl_variable_cost_ub.append(xl_target_data[i]["variable_cost_ub"])
        xl_variable_cost_mean.append(xl_target_data[i]["variable_cost_mean"])
        xl_variable_cost_std.append(xl_target_data[i]["variable_cost_std"])
        xl_variable_cost_high4.append(xl_target_data[i]["variable_cost_high4"])

        xl_cut3_percent.append(xl_target_data[i]["cut3_percent"])
        # =======  results ======
        xl_obj1.append(xl_target_data[i]["obj1"])  # 总性价比
        xl_obj2.append(xl_target_data[i]["obj2"])  # 总任务开始时间和
        xl_obj3.append(xl_target_data[i]["obj3"])  # 总利润
        xl_obj.append(xl_target_data[i]["obj"])  # 函数目标值

        xl_runtime.append(xl_target_data[i]["runtime"])

    df_target_data = {
        '序号': xl_id,

        '有效不等式选择': xl_valid_selection,
        '求解选择': xl_opt_trigger,
        'random_seed': xl_random_seed,
        'gap': xl_gap,
        'scale_phase': xl_scale_phase,
        'scale_stage': xl_scale_stage,
        'size_signature': xl_size_signature,

        '客户数': xl_customer_num,
        '服务数': xl_service_num,
        '产品数': xl_product_num,
        '服务团队数': xl_team_num,
        'base_num_vars': xl_base_num_vars,
        'base_num_constrs': xl_base_num_constrs,
        'root_num_constrs': xl_root_num_constrs,

        '服务时间': xl_service_time,
        '服务价格': xl_service_price,
        '服务成本': xl_service_cost,

        '利润': xl_profit,
        '利润均值': xl_profit_mean,
        '利润标准差': xl_profit_std,
        '利润上四分位': xl_profit_high4,
        '利润下四分位': xl_profit_low4,

        '服务时间下界': xl_service_time_lb,
        '服务价格下界': xl_service_price_lb,
        '服务成本下界': xl_service_cost_lb,

        '服务时间梯度': xl_service_time_step,
        '服务价格梯度': xl_service_price_step,
        '服务成本梯度': xl_service_cost_step,

        '客户效用值': xl_utility,
        '客户效用值下界': xl_utility_lb,
        '客户效用值梯度': xl_utility_step,
        '库存': xl_inventory,
        '库存梯度': xl_inventory_step,
        '库存下界': xl_inventory_lb,

        '产品服务对应关系矩阵': xl_Ser_Pro_corr,
        '产品服务对应关系': xl_Ser_Pro_coe,
        '产品服务对应关系下界': xl_Ser_Pro_lb,
        '产品服务对应关系上界': xl_Ser_Pro_ub,
        # mean
        '产品服务对应关系均值': xl_Ser_Pro_coe_mean,
        # std
        '产品服务对应关系标准差': xl_Ser_Pro_coe_std,
        # high 4
        '产品服务对应关系上四分位': xl_Ser_Pro_coe_high4,

        '客户权重': xl_weight,
        '客户权重gap': xl_weight_gap,
        '客户权重均值': xl_weight_mean,
        '客户权重标准差': xl_weight_std,
        '客户权重上四分位': xl_weight_high4,

        '服务时间窗Early': xl_early_start,
        '服务时间窗Early下界': xl_early_start_lb,
        '服务时间窗Early上界': xl_early_start_ub,
        '服务时间窗Late': xl_late_start,
        '服务时间窗Late下界': xl_late_start_lb,
        '服务时间窗Late上界': xl_late_start_ub,

        # mean
        '服务时间窗Early均值': xl_early_start_mean,
        '服务时间窗Late均值': xl_late_start_mean,
        # std
        '服务时间窗Early标准差': xl_early_start_std,
        '服务时间窗Late标准差': xl_late_start_std,
        # high4
        '服务时间窗Early上四分位': xl_early_start_high4,
        '服务时间窗Late上四分位': xl_late_start_high4,

        '转移时间矩阵': xl_transfer_matrix,
        '转移时间矩阵下界': xl_transfer_lb,
        '转移时间矩阵上界': xl_transfer_ub,
        '转移时间矩阵均值': xl_transfer_mean,
        '转移时间矩阵标准差': xl_transfer_std,
        '转移时间矩阵上四分位': xl_transfer_high4,

        '工作时长': xl_duration,

        '固定成本': xl_fixed_cost,
        '可变成本': xl_variable_cost,
        '可变成本下界': xl_variable_cost_lb,
        '可变成本上界': xl_variable_cost_ub,
        '可变成本均值': xl_variable_cost_mean,
        '可变成本标准差': xl_variable_cost_std,
        '可变成本上四分位': xl_variable_cost_high4,

        'cut3百分比': xl_cut3_percent,
        # =======  results ======
        '总性价比': xl_obj1,
        '总任务开始时间和': xl_obj2,
        '总利润': xl_obj3,
        '函数目标值': xl_obj,

        '求解时间': xl_runtime
    }
    current_time = datetime.datetime.now().strftime('%Y-%m-%d')
    filename =  'result.csv'
    df_target = pd.DataFrame(df_target_data)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    df_target.to_csv(RESULT_CSV_PATH, index=False, mode='a')
    # path = RESULT_CSV_PATH
    # # 使用openpyxl加载现有的Excel文件
    # book = load_workbook(path)
    # # 创建一个pandas Excel writer，使用openpyxl作为引擎
    # writer = pd.ExcelWriter(path, engine='openpyxl')
    # # 将book作为现有的工作簿
    # writer.book = book
    # # 将数据写入Excel
    # df_target.to_excel(writer, index=False, sheet_name='Sheet1', startrow=book['Sheet1'].max_row)
    # # 保存文件
    # writer.save()
    # train = pd.concat([train, current_df])
    print(f'-----------------------------------------------times = {times}')
    times += 1

